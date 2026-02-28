#!/usr/bin/env python3
"""
excel_to_fmd.py — Convert Excel workbooks to Formula Markdown (FMD) format.

Usage:
    python excel_to_fmd.py <path_to_excel_file>

Output:
    A folder named after the Excel file, containing one .fmd.md file per sheet.
    Comments from each sheet are collected under a ## Comments section at the end.
"""

import sys
import os
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


# ---------------------------------------------------------------------------
# Formula translation
# ---------------------------------------------------------------------------

def excel_formula_to_fmd(
    formula: str,
    headers: list[str],
    row_idx: int,
    fallback_value=None,
) -> str:
    """
    Translate an Excel formula to FMD syntax.

    - Same-row cell references        → {Column Header}
    - Single-column aggregate ranges  → FUNC(Column Header)
    - Anything else (cross-row, cross-sheet, absolute, complex)
      → replaced with its computed value (fallback_value).
      If fallback_value is None, the raw Excel ref is kept as a comment.
    """
    if not formula.startswith("="):
        return formula

    expr = formula[1:]

    col_map = {get_column_letter(i + 1): h for i, h in enumerate(headers)}
    untranslatable = []

    def replace_cell_ref(m):
        col_letters = re.sub(r"\$", "", m.group(1))
        row_num = re.sub(r"\$", "", m.group(2))
        if int(row_num) == row_idx and col_letters in col_map:
            return "{" + col_map[col_letters] + "}"
        # Cannot translate — mark it
        untranslatable.append(m.group(0))
        return f"__UNRESOLVABLE__"

    expr = re.sub(r"(\$?[A-Z]{1,3})(\$?\d+)", replace_cell_ref, expr)

    def replace_range(m):
        func = m.group(1).upper()
        range_str = m.group(2)
        range_match = re.match(
            r"\$?([A-Z]{1,3})\$?\d+:\$?([A-Z]{1,3})\$?\d+", range_str
        )
        if range_match and range_match.group(1) == range_match.group(2):
            col = range_match.group(1)
            if col in col_map:
                return f"{func}({col_map[col]})"
        untranslatable.append(m.group(0))
        return f"__UNRESOLVABLE__"

    expr = re.sub(r"([A-Z]+)\(([^)]+)\)", replace_range, expr)

    # If any part couldn't be translated, fall back to the computed value
    if "__UNRESOLVABLE__" in expr or untranslatable:
        if fallback_value is not None:
            return str(fallback_value)
        return formula  # last resort: original Excel formula as plain text

    return "={" + expr + "}"


def make_fmd_formula(
    formula: str,
    headers: list[str],
    row_idx: int,
    fallback_value=None,
) -> str:
    """Return a backtick-wrapped FMD formula, or a plain value if not translatable."""
    result = excel_formula_to_fmd(formula, headers, row_idx, fallback_value)
    # If it translated cleanly it starts with ={
    if result.startswith("={"):
        return f"`{result}`"
    # Otherwise it's a resolved scalar value — return as-is
    return result


# ---------------------------------------------------------------------------
# Markdown table helpers
# ---------------------------------------------------------------------------

def col_widths(rows: list[list[str]]) -> list[int]:
    if not rows:
        return []
    widths = [0] * len(rows[0])
    for row in rows:
        for i, cell in enumerate(row):
            if i < len(widths):
                widths[i] = max(widths[i], len(str(cell)))
    return [max(w, 3) for w in widths]


def format_table(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    widths = col_widths(rows)
    lines = []
    for i, row in enumerate(rows):
        padded = [str(cell).ljust(widths[j]) for j, cell in enumerate(row)]
        lines.append("| " + " | ".join(padded) + " |")
        if i == 0:
            lines.append("|" + "|".join("-" * (w + 2) for w in widths) + "|")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Sheet → FMD conversion
# ---------------------------------------------------------------------------

def cell_value_to_str(cell) -> str:
    """Return the display value of a cell, converting formulas to FMD syntax."""
    return str(cell.value) if cell.value is not None else ""


def sheet_to_fmd(ws, value_ws) -> str:
    """
    Convert a single worksheet to FMD content.

    ws       — workbook sheet loaded with data_only=False (contains formulas)
    value_ws — same sheet loaded with data_only=True  (contains computed values)
    """
    sections = []
    comments = []  # (cell_ref, author, text)

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    if max_row == 0 or max_col == 0:
        return "_Empty sheet._\n"

    # Harvest comments (only available on formula workbook)
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.comment:
                ref = f"{get_column_letter(cell.column)}{cell.row}"
                author = cell.comment.author or "Unknown"
                text = cell.comment.text.strip() if cell.comment.text else ""
                comments.append((ref, author, text))

    # Load both formula rows and value rows together
    formula_rows = list(ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col))
    value_rows   = list(value_ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col))

    headers = [cell_value_to_str(c) for c in formula_rows[0]]

    is_vars_block = (
        max_col == 2
        and len(formula_rows) >= 2
        and str(headers[0]).strip().lower() in ("var", "variable", "name", "key", "parameter")
        and str(headers[1]).strip().lower() in ("value", "val", "default")
    )

    if is_vars_block:
        sections.append("<!-- fmd:vars -->")
        table_rows = [headers]
        for f_row, v_row in zip(formula_rows[1:], value_rows[1:]):
            table_rows.append([
                cell_value_to_str(f) if not str(f.value or "").startswith("=")
                else str(v.value) if v.value is not None
                else ""
                for f, v in zip(f_row, v_row)
            ])
        sections.append(format_table(table_rows))
    else:
        sections.append("<!-- fmd:table -->")
        table_rows = [headers]
        for r_idx, (f_row, v_row) in enumerate(
            zip(formula_rows[1:], value_rows[1:]), start=2
        ):
            display_row = []
            for f_cell, v_cell in zip(f_row, v_row):
                formula = str(f_cell.value) if f_cell.value is not None else ""
                if formula.startswith("="):
                    fallback = v_cell.value  # computed value from data_only wb
                    display_row.append(
                        make_fmd_formula(formula, headers, r_idx, fallback_value=fallback)
                    )
                else:
                    display_row.append(cell_value_to_str(f_cell))
            table_rows.append(display_row)
        sections.append(format_table(table_rows))

    # --- Comments section ---
    if comments:
        sections.append("\n## Comments\n")
        for ref, author, text in comments:
            sections.append(f"**{ref}** _(by {author})_: {text}\n")

    return "\n\n".join(sections) + "\n"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def excel_to_fmd(excel_path: str):
    path = Path(excel_path)
    if not path.exists():
        print(f"Error: File not found — {excel_path}")
        sys.exit(1)

    out_dir = Path(path.stem)
    out_dir.mkdir(exist_ok=True)

    print(f"Loading workbook: {path.name}")
    wb_formulas = load_workbook(path, data_only=False)  # formulas as strings
    wb_values   = load_workbook(path, data_only=True)   # computed values

    for sheet_name in wb_formulas.sheetnames:
        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name]
        print(f"  Processing sheet: {sheet_name}")

        fmd_content = f"# {sheet_name}\n\n" + sheet_to_fmd(ws_f, ws_v)

        safe_name = re.sub(r'[\\/*?:"<>|]', "_", sheet_name)
        out_file = out_dir / f"{safe_name}.md"

        out_file.write_text(fmd_content, encoding="utf-8")
        print(f"    → {out_file}")

    print(f"\nDone. Output folder: {out_dir}/")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python excel_to_fmd.py <path_to_excel_file>")
        sys.exit(1)
    excel_to_fmd(sys.argv[1])
